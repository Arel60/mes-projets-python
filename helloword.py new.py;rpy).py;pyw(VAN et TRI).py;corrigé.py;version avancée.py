
import math
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime

import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# OUTILS DE FORMATAGE
# ============================================================

def formater_nombre(nombre, decimales=2):
    """Formate un nombre avec un affichage adapté au français."""
    if nombre is None:
        return "-"

    try:
        texte = f"{nombre:,.{decimales}f}"
        return texte.replace(",", " ").replace(".", ",")
    except (ValueError, TypeError):
        return "-"


def formater_pourcentage(nombre, decimales=2):
    """Transforme un taux décimal en pourcentage."""
    if nombre is None:
        return "-"

    return f"{nombre * 100:.{decimales}f}".replace(".", ",") + " %"


def convertir_nombre(texte):
    """
    Convertit un nombre saisi par l'utilisateur.
    Accepte :
        1000
        1000.50
        1000,50
        1 000,50
        1.000,50
    """
    texte = str(texte).strip()

    if not texte:
        raise ValueError("Une valeur est obligatoire.")

    texte = texte.replace("\u00a0", " ")
    texte = texte.replace(" ", "")

    # Gestion du format français 1.000,50
    if "," in texte and "." in texte:
        if texte.rfind(",") > texte.rfind("."):
            texte = texte.replace(".", "").replace(",", ".")
        else:
            texte = texte.replace(",", "")
    else:
        texte = texte.replace(",", ".")

    try:
        valeur = float(texte)
    except ValueError:
        raise ValueError(f"Valeur numérique invalide : {texte}")

    if not math.isfinite(valeur):
        raise ValueError("La valeur saisie doit être un nombre fini.")

    return valeur


# ============================================================
# 1. CALCUL DE LA VAN
# ============================================================

def calcul_van(flux, taux):
    """
    Calcule la Valeur Actuelle Nette.

    Le flux[0] correspond à l'investissement initial à la date 0.
    """
    if not isinstance(flux, (list, tuple)) or len(flux) == 0:
        raise ValueError("Les flux doivent contenir au moins une valeur.")

    if not math.isfinite(taux):
        raise ValueError("Le taux d'actualisation doit être un nombre fini.")

    if taux <= -1:
        raise ValueError(
            "Le taux d'actualisation doit être supérieur à -100 %."
        )

    van = 0.0

    for annee, valeur in enumerate(flux):
        if not isinstance(valeur, (int, float)) or not math.isfinite(valeur):
            raise ValueError("Tous les flux doivent être des nombres réels finis.")

        van += valeur / ((1 + taux) ** annee)

    return van


# ============================================================
# 2. CALCUL DU TRI
# ============================================================

def trouver_racines_tri(flux, precision=1e-8):
    """
    Recherche les différents TRI possibles.

    Une série de flux peut posséder plusieurs TRI lorsqu'elle comporte
    plusieurs changements de signe.

    Retourne une liste de TRI.
    """
    if not isinstance(flux, (list, tuple)):
        raise ValueError("Les flux doivent être fournis sous forme de liste.")

    if len(flux) < 2:
        raise ValueError("Il faut au moins deux flux pour calculer un TRI.")

    if not all(
        isinstance(f, (int, float)) and math.isfinite(f)
        for f in flux
    ):
        raise ValueError(
            "Tous les flux doivent être des nombres réels finis."
        )

    if not any(f < 0 for f in flux):
        raise ValueError("Aucun flux négatif n'a été détecté.")

    if not any(f > 0 for f in flux):
        raise ValueError("Aucun flux positif n'a été détecté.")

    # --------------------------------------------------------
    # Création d'une grille de recherche.
    # On commence juste au-dessus de -100 %.
    # --------------------------------------------------------

    taux_min = -0.999999
    taux_max = 1.0

    valeurs_taux = []

    # Zone dense autour des taux courants.
    nombre_points = 500

    for i in range(nombre_points + 1):
        taux = taux_min + (
            (taux_max - taux_min) * i / nombre_points
        )
        valeurs_taux.append(taux)

    # Extension progressive jusqu'à un taux très élevé.
    taux = taux_max

    for _ in range(12):
        taux *= 2
        valeurs_taux.append(taux)

    valeurs_taux = sorted(set(valeurs_taux))

    racines = []

    van_precedente = calcul_van(flux, valeurs_taux[0])

    # --------------------------------------------------------
    # Recherche des intervalles contenant une racine.
    # --------------------------------------------------------

    for i in range(1, len(valeurs_taux)):
        taux_gauche = valeurs_taux[i - 1]
        taux_droit = valeurs_taux[i]

        van_gauche = van_precedente
        van_droit = calcul_van(flux, taux_droit)

        # Racine exacte
        if abs(van_gauche) <= precision:
            racines.append(taux_gauche)

        # Changement de signe
        elif van_gauche * van_droit < 0:
            gauche = taux_gauche
            droite = taux_droit
            van_g = van_gauche

            iterations = 0

            while droite - gauche > precision:
                milieu = (gauche + droite) / 2
                van_m = calcul_van(flux, milieu)

                if abs(van_m) <= precision:
                    gauche = milieu
                    droite = milieu
                    break

                if van_g * van_m <= 0:
                    droite = milieu
                else:
                    gauche = milieu
                    van_g = van_m

                iterations += 1

                if iterations > 5000:
                    break

            racine = (gauche + droite) / 2

            if math.isfinite(racine):
                racines.append(racine)

        van_precedente = van_droit

    # Vérification du dernier point
    if abs(van_precedente) <= precision:
        racines.append(valeurs_taux[-1])

    # Suppression des doublons
    racines_uniques = []

    for racine in sorted(racines):
        if not racines_uniques:
            racines_uniques.append(racine)
        elif abs(racine - racines_uniques[-1]) > precision * 10:
            racines_uniques.append(racine)

    if not racines_uniques:
        raise ValueError(
            "Impossible de déterminer le TRI avec les flux fournis."
        )

    return racines_uniques


def calcul_tri(flux, precision=1e-8):
    """
    Calcule le TRI principal.

    Retourne :
        tri,
        nombre de changements de signe,
        liste des TRI détectés
    """
    if precision <= 0 or precision >= 1:
        raise ValueError(
            "La précision doit être comprise entre 0 et 1."
        )

    changements_signe = 0

    for i in range(len(flux) - 1):
        if flux[i] * flux[i + 1] < 0:
            changements_signe += 1

    tris = trouver_racines_tri(flux, precision)

    # Le premier TRI détecté est retenu comme TRI principal.
    tri = tris[0]

    return tri, changements_signe, tris


# ============================================================
# 3. ANALYSE DU PROJET
# ============================================================

def analyser_projet(flux, taux, numero):
    """Analyse financière complète d'un projet."""

    if len(flux) < 2:
        raise ValueError("Le projet doit comporter au moins un flux annuel.")

    investissement = abs(flux[0])

    if investissement <= 0:
        raise ValueError(
            "L'investissement initial doit être supérieur à zéro."
        )

    tri, changements_signe, tris = calcul_tri(flux)

    van = calcul_van(flux, taux)

    # --------------------------------------------------------
    # Indice de rentabilité
    # --------------------------------------------------------

    valeur_actuelle_flux_futurs = van + investissement

    indice = valeur_actuelle_flux_futurs / investissement

    # --------------------------------------------------------
    # Délai de récupération simple
    # --------------------------------------------------------

    cumul = flux[0]
    recuperation = None

    for annee in range(1, len(flux)):
        cumul_precedent = cumul
        cumul += flux[annee]

        if cumul >= 0:
            if flux[annee] != 0:
                fraction = -cumul_precedent / flux[annee]
                recuperation = (annee - 1) + fraction
            else:
                recuperation = float(annee)

            break

    # --------------------------------------------------------
    # Décision
    # --------------------------------------------------------

    if van > 0 and tri > taux:
        decision = "RENTABLE ET ACCEPTABLE"

    elif van < 0 and tri < taux:
        decision = "NON RENTABLE - À REJETER"

    elif abs(van) < 1e-9:
        decision = "ÉQUILIBRE - VAN NULLE"

    else:
        decision = "À ANALYSER AVEC PRUDENCE"

    # --------------------------------------------------------
    # Vérification du TRI
    # --------------------------------------------------------

    van_verification = calcul_van(flux, tri)

    tri_precis = abs(van_verification) <= 0.01

    return {
        "numero": numero,
        "investissement": investissement,
        "annees": len(flux) - 1,
        "taux": taux,
        "van": van,
        "tri": tri,
        "tris": tris,
        "indice": indice,
        "recuperation": recuperation,
        "flux": list(flux),
        "decision": decision,
        "van_verification": van_verification,
        "tri_precis": tri_precis,
        "changements_signe": changements_signe
    }


# ============================================================
# 4. APPLICATION GRAPHIQUE
# ============================================================

class CalculateurFinancier:

    def __init__(self, fenetre):

        self.fenetre = fenetre

        self.fenetre.title(
            "Calculateur Financier Professionnel - VAN & TRI"
        )

        self.fenetre.geometry("1350x900")
        self.fenetre.minsize(1150, 750)

        self.projets = []
        self.numero_projet = 1
        self.entrees_flux = []

        self.creer_styles()
        self.creer_interface()

    # ========================================================
    # STYLES
    # ========================================================

    def creer_styles(self):

        style = ttk.Style()

        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(
            "Treeview",
            font=("Arial", 10),
            rowheight=34
        )

        style.configure(
            "Treeview.Heading",
            font=("Arial", 10, "bold"),
            padding=8
        )

        # Correction de l'erreur de syntaxe présente dans le code initial.
        style.configure(
            "TNotebook.Tab",
            font=("Arial", 10, "bold")
        )

        style.map(
            "Treeview",
            background=[
                ("selected", "#D9EAF7")
            ],
            foreground=[
                ("selected", "black")
            ]
        )

    # ========================================================
    # INTERFACE
    # ========================================================

    def creer_interface(self):

        titre = tk.Label(
            self.fenetre,
            text="CALCULATEUR FINANCIER",
            font=("Arial", 24, "bold")
        )

        titre.pack(pady=(15, 2))

        sous_titre = tk.Label(
            self.fenetre,
            text="Analyse professionnelle des projets d'investissement",
            font=("Arial", 12)
        )

        sous_titre.pack(pady=(0, 12))

        # ----------------------------------------------------
        # DONNÉES DU PROJET
        # ----------------------------------------------------

        cadre_donnees = tk.LabelFrame(
            self.fenetre,
            text=" DONNÉES DU PROJET ",
            font=("Arial", 12, "bold"),
            padx=15,
            pady=10
        )

        cadre_donnees.pack(
            fill="x",
            padx=20,
            pady=5
        )

        # Investissement
        tk.Label(
            cadre_donnees,
            text="Investissement initial :",
            font=("Arial", 10)
        ).grid(
            row=0,
            column=0,
            sticky="w",
            padx=5,
            pady=5
        )

        self.entree_investissement = tk.Entry(
            cadre_donnees,
            width=20,
            font=("Arial", 10)
        )

        self.entree_investissement.grid(
            row=0,
            column=1,
            padx=5
        )

        tk.Label(
            cadre_donnees,
            text="FCFA"
        ).grid(
            row=0,
            column=2,
            sticky="w"
        )

        # Années
        tk.Label(
            cadre_donnees,
            text="Nombre d'années :",
            font=("Arial", 10)
        ).grid(
            row=1,
            column=0,
            sticky="w",
            padx=5,
            pady=5
        )

        self.entree_annees = tk.Entry(
            cadre_donnees,
            width=20,
            font=("Arial", 10)
        )

        self.entree_annees.grid(
            row=1,
            column=1,
            padx=5
        )

        # Taux
        tk.Label(
            cadre_donnees,
            text="Taux d'actualisation :",
            font=("Arial", 10)
        ).grid(
            row=2,
            column=0,
            sticky="w",
            padx=5,
            pady=5
        )

        self.entree_taux = tk.Entry(
            cadre_donnees,
            width=20,
            font=("Arial", 10)
        )

        self.entree_taux.grid(
            row=2,
            column=1,
            padx=5
        )

        tk.Label(
            cadre_donnees,
            text="%"
        ).grid(
            row=2,
            column=2,
            sticky="w"
        )

        # Bouton générer
        tk.Button(
            cadre_donnees,
            text="GÉNÉRER LES FLUX",
            command=self.generer_flux,
            width=22,
            height=2,
            font=("Arial", 10, "bold")
        ).grid(
            row=0,
            column=4,
            rowspan=3,
            padx=30
        )

        # ----------------------------------------------------
        # FLUX
        # ----------------------------------------------------

        self.cadre_flux = tk.LabelFrame(
            self.fenetre,
            text=" FLUX FINANCIERS ",
            font=("Arial", 12, "bold"),
            padx=10,
            pady=8
        )

        self.cadre_flux.pack(
            fill="both",
            padx=20,
            pady=5
        )

        tk.Label(
            self.cadre_flux,
            text=(
                "Entrez le nombre d'années puis cliquez sur "
                "« Générer les flux »."
            )
        ).pack()

        # ----------------------------------------------------
        # BOUTONS
        # ----------------------------------------------------

        cadre_boutons = tk.Frame(self.fenetre)
        cadre_boutons.pack(pady=8)

        boutons = [
            ("ANALYSER", self.analyser),
            ("NOUVEAU PROJET", self.nouveau_projet),
            ("SENSIBILITÉ", self.sensibilite),
            ("GRAPHIQUE", self.graphique),
            ("COMPARER", self.comparer),
            ("EXCEL", self.exporter_excel),
            ("RAPPORT", self.generer_rapport)
        ]

        for colonne, (texte, commande) in enumerate(boutons):

            tk.Button(
                cadre_boutons,
                text=texte,
                command=commande,
                width=15,
                height=2,
                font=("Arial", 9, "bold")
            ).grid(
                row=0,
                column=colonne,
                padx=3
            )

        # ----------------------------------------------------
        # TABLEAU
        # ----------------------------------------------------

        cadre_tableau = tk.LabelFrame(
            self.fenetre,
            text=" TABLEAU DE SYNTHÈSE DES PROJETS ",
            font=("Arial", 12, "bold"),
            padx=5,
            pady=5
        )

        cadre_tableau.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=5
        )

        colonnes = (
            "Projet",
            "Investissement",
            "Années",
            "Taux",
            "VAN",
            "TRI",
            "Indice",
            "Récupération",
            "Décision"
        )

        cadre_tree = tk.Frame(cadre_tableau)
        cadre_tree.pack(fill="both", expand=True)

        scrollbar_verticale = ttk.Scrollbar(
            cadre_tree,
            orient="vertical"
        )

        scrollbar_horizontale = ttk.Scrollbar(
            cadre_tree,
            orient="horizontal"
        )

        self.tableau = ttk.Treeview(
            cadre_tree,
            columns=colonnes,
            show="headings",
            yscrollcommand=scrollbar_verticale.set,
            xscrollcommand=scrollbar_horizontale.set,
            selectmode="browse"
        )

        scrollbar_verticale.config(
            command=self.tableau.yview
        )

        scrollbar_horizontale.config(
            command=self.tableau.xview
        )

        scrollbar_verticale.pack(
            side="right",
            fill="y"
        )

        scrollbar_horizontale.pack(
            side="bottom",
            fill="x"
        )

        self.tableau.pack(
            fill="both",
            expand=True
        )

        titres_colonnes = {
            "Projet": "PROJET",
            "Investissement": "INVESTISSEMENT INITIAL",
            "Années": "ANNÉES",
            "Taux": "TAUX (%)",
            "VAN": "VAN (FCFA)",
            "TRI": "TRI (%)",
            "Indice": "INDICE",
            "Récupération": "RÉCUPÉRATION",
            "Décision": "DÉCISION"
        }

        largeurs = {
            "Projet": 90,
            "Investissement": 190,
            "Années": 80,
            "Taux": 100,
            "VAN": 190,
            "TRI": 100,
            "Indice": 90,
            "Récupération": 140,
            "Décision": 280
        }

        for colonne in colonnes:

            self.tableau.heading(
                colonne,
                text=titres_colonnes[colonne]
            )

            self.tableau.column(
                colonne,
                width=largeurs[colonne],
                anchor="center"
            )

        self.tableau.tag_configure(
            "pair",
            background="#F4F6F7"
        )

        self.tableau.tag_configure(
            "impair",
            background="#FFFFFF"
        )

        self.tableau.tag_configure(
            "rentable",
            foreground="#087A28"
        )

        self.tableau.tag_configure(
            "rejeter",
            foreground="#B00020"
        )

        self.tableau.tag_configure(
            "prudence",
            foreground="#9A6700"
        )

        self.tableau.bind(
            "<Double-1>",
            self.details_projet
        )

        # ----------------------------------------------------
        # DÉTAILS
        # ----------------------------------------------------

        cadre_details = tk.LabelFrame(
            self.fenetre,
            text=" DÉTAIL DU PROJET SÉLECTIONNÉ ",
            font=("Arial", 11, "bold"),
            padx=5,
            pady=5
        )

        cadre_details.pack(
            fill="x",
            padx=20,
            pady=(0, 10)
        )

        self.zone_resultats = tk.Text(
            cadre_details,
            height=7,
            font=("Consolas", 10),
            wrap="word"
        )

        self.zone_resultats.pack(fill="x")

        self.zone_resultats.insert(
            tk.END,
            "Sélectionnez un projet dans le tableau pour afficher ses détails."
        )

    # ========================================================
    # GÉNÉRER LES FLUX
    # ========================================================

    def generer_flux(self):

        try:
            annees = int(self.entree_annees.get().strip())

            if annees <= 0:
                raise ValueError(
                    "Le nombre d'années doit être supérieur à 0."
                )

            if annees > 100:
                raise ValueError(
                    "Le nombre d'années ne peut pas dépasser 100."
                )

        except ValueError as erreur:

            messagebox.showerror(
                "Erreur",
                str(erreur)
            )

            return

        for widget in self.cadre_flux.winfo_children():
            widget.destroy()

        self.entrees_flux = []

        tk.Label(
            self.cadre_flux,
            text="ANNÉE",
            font=("Arial", 10, "bold")
        ).grid(
            row=0,
            column=0,
            padx=20,
            pady=5
        )

        tk.Label(
            self.cadre_flux,
            text="FLUX (FCFA)",
            font=("Arial", 10, "bold")
        ).grid(
            row=0,
            column=1,
            padx=20,
            pady=5
        )

        for annee in range(1, annees + 1):

            tk.Label(
                self.cadre_flux,
                text=f"Année {annee}"
            ).grid(
                row=annee,
                column=0,
                padx=20,
                pady=2
            )

            entree = tk.Entry(
                self.cadre_flux,
                width=25,
                font=("Arial", 10)
            )

            entree.grid(
                row=annee,
                column=1,
                padx=20,
                pady=2
            )

            self.entrees_flux.append(entree)

        tk.Label(
            self.cadre_flux,
            text=f"{annees} flux annuels à renseigner."
        ).grid(
            row=annees + 1,
            column=0,
            columnspan=2,
            pady=5
        )

        # Positionner automatiquement le curseur dans le premier flux.
        if self.entrees_flux:
            self.entrees_flux[0].focus_set()

    # ========================================================
    # ANALYSER
    # ========================================================

    def analyser(self):

        try:
            investissement = convertir_nombre(
                self.entree_investissement.get()
            )

            annees = int(
                self.entree_annees.get().strip()
            )

            taux_pourcentage = convertir_nombre(
                self.entree_taux.get()
            )

            if investissement <= 0:
                raise ValueError(
                    "L'investissement doit être supérieur à 0."
                )

            if annees <= 0:
                raise ValueError(
                    "Le nombre d'années doit être supérieur à 0."
                )

            if annees > 100:
                raise ValueError(
                    "Le nombre d'années ne peut pas dépasser 100."
                )

            if taux_pourcentage <= -100:
                raise ValueError(
                    "Le taux doit être supérieur à -100 %."
                )

            if len(self.entrees_flux) != annees:
                raise ValueError(
                    "Cliquez d'abord sur « Générer les flux »."
                )

            taux = taux_pourcentage / 100

            flux = [-investissement]

            for numero_annee, entree in enumerate(
                self.entrees_flux,
                start=1
            ):

                texte = entree.get().strip()

                if texte == "":
                    raise ValueError(
                        f"Le flux de l'année {numero_annee} "
                        "doit être renseigné."
                    )

                valeur = convertir_nombre(texte)

                flux.append(valeur)

            resultat = analyser_projet(
                flux,
                taux,
                self.numero_projet
            )

            self.projets.append(resultat)

            self.ajouter_au_tableau(resultat)
            self.afficher_resultat(resultat)

            messagebox.showinfo(
                "Analyse terminée",
                f"Projet {self.numero_projet} analysé avec succès."
            )

        except ValueError as erreur:

            messagebox.showerror(
                "Erreur de saisie",
                str(erreur)
            )

        except Exception as erreur:

            messagebox.showerror(
                "Erreur inattendue",
                f"Une erreur inattendue est survenue :\n{erreur}"
            )

    # ========================================================
    # AJOUT AU TABLEAU
    # ========================================================

    def ajouter_au_tableau(self, projet):

        investissement = formater_nombre(
            projet["investissement"]
        )

        van = formater_nombre(
            projet["van"]
        )

        taux = formater_pourcentage(
            projet["taux"]
        )

        tri = formater_pourcentage(
            projet["tri"]
        )

        indice = formater_nombre(
            projet["indice"]
        )

        if projet["recuperation"] is not None:

            recuperation = (
                formater_nombre(
                    projet["recuperation"]
                )
                + " ans"
            )

        else:

            recuperation = "Non récupéré"

        if "RENTABLE" in projet["decision"]:

            tag_decision = "rentable"

        elif "REJETER" in projet["decision"]:

            tag_decision = "rejeter"

        else:

            tag_decision = "prudence"

        tag_ligne = (
            "pair"
            if len(self.tableau.get_children()) % 2 == 0
            else "impair"
        )

        self.tableau.insert(
            "",
            tk.END,
            values=(
                f"Projet {projet['numero']}",
                investissement + " FCFA",
                projet["annees"],
                taux,
                van + " FCFA",
                tri,
                indice,
                recuperation,
                projet["decision"]
            ),
            tags=(tag_ligne, tag_decision)
        )

    # ========================================================
    # AFFICHER RÉSULTAT
    # ========================================================

    def afficher_resultat(self, projet):

        self.zone_resultats.delete(
            "1.0",
            tk.END
        )

        van = formater_nombre(
            projet["van"]
        )

        tri = formater_pourcentage(
            projet["tri"]
        )

        indice = formater_nombre(
            projet["indice"]
        )

        texte = (
            f"PROJET {projet['numero']}\n"
            f"{'=' * 70}\n"
            f"Investissement initial : "
            f"{formater_nombre(projet['investissement'])} FCFA\n"
            f"Durée : {projet['annees']} ans\n"
            f"Taux d'actualisation : "
            f"{formater_pourcentage(projet['taux'])}\n"
            f"VAN : {van} FCFA\n"
            f"TRI : {tri}\n"
            f"Indice de rentabilité : {indice}\n"
        )

        if projet["recuperation"] is not None:

            recuperation = formater_nombre(
                projet["recuperation"]
            )

            texte += (
                f"Délai de récupération : "
                f"{recuperation} ans\n"
            )

        else:

            texte += (
                "Délai de récupération : Non récupéré\n"
            )

        texte += (
            f"Décision : {projet['decision']}\n"
            f"VAN au TRI : "
            f"{projet['van_verification']:.8f}\n"
        )

        if projet["tri_precis"]:

            texte += (
                "Vérification : TRI suffisamment précis.\n"
            )

        else:

            texte += (
                "Vérification : précision du TRI à contrôler.\n"
            )

        if len(projet["tris"]) > 1:

            texte += (
                "\nATTENTION : plusieurs TRI ont été détectés.\n"
                f"TRI détectés : "
                f"{', '.join(formater_pourcentage(t) for t in projet['tris'])}\n"
            )

        elif projet["changements_signe"] > 1:

            texte += (
                "\nATTENTION : plusieurs changements de signe "
                "ont été détectés.\n"
                "Le projet peut potentiellement posséder plusieurs TRI."
            )

        self.zone_resultats.insert(
            tk.END,
            texte
        )

    # ========================================================
    # DOUBLE-CLIC
    # ========================================================

    def details_projet(self, evenement=None):

        selection = self.tableau.selection()

        if not selection:
            return

        element = selection[0]

        valeurs = self.tableau.item(
            element,
            "values"
        )

        if not valeurs:
            return

        try:

            numero = int(
                str(valeurs[0]).split()[-1]
            )

        except (ValueError, IndexError):
            return

        for projet in self.projets:

            if projet["numero"] == numero:

                self.afficher_resultat(projet)
                break

    # ========================================================
    # NOUVEAU PROJET
    # ========================================================

    def nouveau_projet(self):

        donnees_presentes = (
            self.entree_investissement.get().strip()
            or self.entree_annees.get().strip()
            or self.entree_taux.get().strip()
            or self.entrees_flux
        )

        if donnees_presentes:

            confirmation = messagebox.askyesno(
                "Nouveau projet",
                "Voulez-vous vraiment commencer un nouveau projet ?"
            )

            if not confirmation:
                return

        self.numero_projet += 1

        self.entree_investissement.delete(0, tk.END)
        self.entree_annees.delete(0, tk.END)
        self.entree_taux.delete(0, tk.END)

        for widget in self.cadre_flux.winfo_children():
            widget.destroy()

        self.entrees_flux = []

        tk.Label(
            self.cadre_flux,
            text=(
                "Entrez le nombre d'années puis cliquez sur "
                "« Générer les flux »."
            )
        ).pack()

        self.zone_resultats.delete(
            "1.0",
            tk.END
        )

        self.zone_resultats.insert(
            tk.END,
            f"Nouveau projet prêt : Projet {self.numero_projet}."
        )

        self.entree_investissement.focus_set()

    # ========================================================
    # SENSIBILITÉ
    # ========================================================

    def sensibilite(self):

        if not self.projets:

            messagebox.showwarning(
                "Attention",
                "Analysez d'abord un projet."
            )

            return

        projet = self.projets[-1]

        fenetre = tk.Toplevel(self.fenetre)

        fenetre.title("Analyse de sensibilité")
        fenetre.geometry("700x600")
        fenetre.transient(self.fenetre)

        tk.Label(
            fenetre,
            text="ANALYSE DE SENSIBILITÉ",
            font=("Arial", 18, "bold")
        ).pack(pady=15)

        tk.Label(
            fenetre,
            text=(
                f"Projet {projet['numero']} — "
                "Impact du taux sur la VAN"
            ),
            font=("Arial", 10)
        ).pack()

        cadre = tk.Frame(fenetre)
        cadre.pack(pady=15)

        champs = [
            ("Taux minimum (%) :", "0"),
            ("Taux maximum (%) :", "50"),
            ("Pas (%) :", "5")
        ]

        entrees = []

        for ligne, (texte, valeur) in enumerate(champs):

            tk.Label(
                cadre,
                text=texte
            ).grid(
                row=ligne,
                column=0,
                padx=10,
                pady=5
            )

            entree = tk.Entry(
                cadre,
                width=15
            )

            entree.insert(0, valeur)

            entree.grid(
                row=ligne,
                column=1
            )

            entrees.append(entree)

        entree_min = entrees[0]
        entree_max = entrees[1]
        entree_pas = entrees[2]

        zone = tk.Text(
            fenetre,
            font=("Consolas", 10)
        )

        zone.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=10
        )

        def calculer():

            try:

                taux_min = convertir_nombre(
                    entree_min.get()
                )

                taux_max = convertir_nombre(
                    entree_max.get()
                )

                pas = convertir_nombre(
                    entree_pas.get()
                )

                if taux_min <= -100:
                    raise ValueError(
                        "Le taux minimum doit être supérieur à -100 %."
                    )

                if taux_max < taux_min:
                    raise ValueError(
                        "Le taux maximum doit être supérieur "
                        "au taux minimum."
                    )

                if pas <= 0:
                    raise ValueError(
                        "Le pas doit être supérieur à 0."
                    )

                nombre_points = (
                    int((taux_max - taux_min) / pas) + 1
                )

                if nombre_points > 5000:
                    raise ValueError(
                        "Le pas est trop petit pour cette plage. "
                        "Limitez l'analyse à 5 000 points maximum."
                    )

                zone.delete(
                    "1.0",
                    tk.END
                )

                zone.insert(
                    tk.END,
                    f"{'TAUX':<15}{'VAN (FCFA)':>25}\n"
                )

                zone.insert(
                    tk.END,
                    "-" * 40 + "\n"
                )

                taux_actuel = taux_min

                for _ in range(nombre_points):

                    if taux_actuel > taux_max + 1e-10:
                        break

                    van = calcul_van(
                        projet["flux"],
                        taux_actuel / 100
                    )

                    van_formatee = formater_nombre(van)

                    zone.insert(
                        tk.END,
                        f"{taux_actuel:>7.2f} %"
                        f"{van_formatee:>25}\n"
                    )

                    taux_actuel += pas

            except ValueError as erreur:

                messagebox.showerror(
                    "Erreur",
                    str(erreur),
                    parent=fenetre
                )

        tk.Button(
            fenetre,
            text="CALCULER",
            command=calculer,
            width=20,
            font=("Arial", 10, "bold")
        ).pack(pady=10)

    # ========================================================
    # GRAPHIQUE
    # ========================================================

    def graphique(self):

        if not self.projets:

            messagebox.showwarning(
                "Attention",
                "Analysez d'abord un projet."
            )

            return

        projet = self.projets[-1]

        taux_graphique = []
        van_graphique = []

        # Graphique de 0 % à 100 %.
        for i in range(201):

            taux = i / 200

            van = calcul_van(
                projet["flux"],
                taux
            )

            taux_graphique.append(
                taux * 100
            )

            van_graphique.append(van)

        plt.figure(
            figsize=(10, 6)
        )

        plt.plot(
            taux_graphique,
            van_graphique,
            color="#1f77b4",
            linewidth=2,
            label="VAN"
        )

        plt.axhline(
            y=0,
            color="black",
            linestyle="--",
            linewidth=1
        )

        tri_pourcentage = projet["tri"] * 100

        if 0 <= tri_pourcentage <= 100:

            plt.axvline(
                x=tri_pourcentage,
                color="red",
                linestyle="--",
                linewidth=1.5,
                label=f"TRI = {tri_pourcentage:.2f} %"
            )

        plt.xlabel(
            "Taux d'actualisation (%)"
        )

        plt.ylabel(
            "VAN (FCFA)"
        )

        plt.title(
            f"Évolution de la VAN — Projet {projet['numero']}"
        )

        plt.grid(
            True,
            alpha=0.3
        )

        plt.legend()
        plt.tight_layout()
        plt.show()

    # ========================================================
    # COMPARAISON
    # ========================================================

    def comparer(self):

        if len(self.projets) < 2:

            messagebox.showwarning(
                "Attention",
                "Il faut analyser au moins deux projets."
            )

            return

        fenetre = tk.Toplevel(self.fenetre)

        fenetre.title("Comparaison des projets")
        fenetre.geometry("1100x600")
        fenetre.transient(self.fenetre)

        tk.Label(
            fenetre,
            text="COMPARAISON DES PROJETS",
            font=("Arial", 18, "bold")
        ).pack(pady=15)

        colonnes = (
            "Projet",
            "VAN",
            "TRI",
            "Indice",
            "Récupération",
            "Décision"
        )

        cadre = tk.Frame(fenetre)
        cadre.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=10
        )

        tableau = ttk.Treeview(
            cadre,
            columns=colonnes,
            show="headings"
        )

        scrollbar = ttk.Scrollbar(
            cadre,
            orient="vertical",
            command=tableau.yview
        )

        tableau.configure(
            yscrollcommand=scrollbar.set
        )

        scrollbar.pack(
            side="right",
            fill="y"
        )

        tableau.pack(
            side="left",
            fill="both",
            expand=True
        )

        largeurs = {
            "Projet": 100,
            "VAN": 180,
            "TRI": 120,
            "Indice": 120,
            "Récupération": 180,
            "Décision": 300
        }

        for colonne in colonnes:

            tableau.heading(
                colonne,
                text=colonne
            )

            tableau.column(
                colonne,
                width=largeurs[colonne],
                anchor="center"
            )

        for projet in self.projets:

            van = (
                formater_nombre(projet["van"])
                + " FCFA")

            tri = formater_pourcentage(
                projet["tri"])

            indice = formater_nombre(
                projet["indice"])

            if projet["recuperation"] is not None:

                recuperation = (
                    formater_nombre(
                        projet["recuperation"]
                    )
                    + " ans"
                )

            else:

                recuperation = "Non récupéré"

            tableau.insert(
                "",
                tk.END,
                values=(
                    f"Projet {projet['numero']}",
                    van,
                    tri,
                    indice,
                    recuperation,
                    projet["decision"]
                )
            )

        classement = sorted(
            self.projets,
            key=lambda p: p["van"],
            reverse=True
        )

        meilleur = classement[0]

        tk.Label(
            fenetre,
            text=(
                f"Projet recommandé selon la VAN : "
                f"Projet {meilleur['numero']}"
            ),
            font=("Arial", 12, "bold"),
            fg="#087A28"
        ).pack(pady=10)

    # ========================================================
    # EXPORT EXCEL PROFESSIONNEL
    # ========================================================

    def exporter_excel(self):

        if not self.projets:

            messagebox.showwarning(
                "Attention",
                "Aucun projet à exporter."
            )

            return

        fichier = filedialog.asksaveasfilename(
            title="Enregistrer le fichier Excel",
            defaultextension=".xlsx",
            filetypes=[
                ("Fichier Excel", "*.xlsx")
            ],
            initialfile="resultats_projets.xlsx"
        )

        if not fichier:
            return

        try:

            classeur = Workbook()

            feuille = classeur.active
            feuille.title = "Résultats"

            # ------------------------------------------------
            # COULEURS
            # ------------------------------------------------

            couleur_titre = PatternFill(
                "solid",
                fgColor="1F4E78"
            )

            couleur_entete = PatternFill(
                "solid",
                fgColor="D9EAF7"
            )

            # ------------------------------------------------
            # TITRE
            # ------------------------------------------------

            feuille["A1"] = (
                "CALCULATEUR FINANCIER — VAN & TRI"
            )

            feuille["A1"].font = Font(
                bold=True,
                size=16,
                color="FFFFFF"
            )

            feuille["A1"].fill = couleur_titre

            feuille["A1"].alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            feuille.merge_cells("A1:I1")

            feuille.row_dimensions[1].height = 28

            feuille["A2"] = (
                "Rapport automatique des projets analysés"
            )

            feuille.merge_cells("A2:I2")

            feuille["A2"].alignment = Alignment(
                horizontal="center"
            )

            # ------------------------------------------------
            # EN-TÊTES
            # ------------------------------------------------

            entetes = [
                "Projet",
                "Investissement initial",
                "Années",
                "Taux actualisation (%)",
                "VAN",
                "TRI (%)",
                "Indice de rentabilité",
                "Délai récupération",
                "Décision"
            ]

            for colonne, valeur in enumerate(entetes, 1):

                cellule = feuille.cell(
                    row=4,
                    column=colonne,
                    value=valeur
                )

                cellule.font = Font(
                    bold=True
                )

                cellule.fill = couleur_entete

                cellule.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

            # ------------------------------------------------
            # DONNÉES
            # ------------------------------------------------

            ligne = 5

            for projet in self.projets:

                valeurs = [
                    projet["numero"],
                    projet["investissement"],
                    projet["annees"],
                    projet["taux"] * 100,
                    projet["van"],
                    projet["tri"] * 100,
                    projet["indice"],
                    projet["recuperation"],
                    projet["decision"]
                ]

                for colonne, valeur in enumerate(
                    valeurs,
                    1
                ):

                    cellule = feuille.cell(
                        row=ligne,
                        column=colonne,
                        value=valeur
                    )

                    cellule.alignment = Alignment(
                        vertical="center",
                        wrap_text=True
                    )

                ligne += 1

            # ------------------------------------------------
            # FORMATS NUMÉRIQUES
            # ------------------------------------------------

            for row in range(5, ligne):

                feuille.cell(
                    row=row,
                    column=2
                ).number_format = '#,##0.00'

                feuille.cell(
                    row=row,
                    column=4
                ).number_format = '0.00'

                feuille.cell(
                    row=row,
                    column=5
                ).number_format = '#,##0.00'

                feuille.cell(
                    row=row,
                    column=6
                ).number_format = '0.00'

                feuille.cell(
                    row=row,
                    column=7
                ).number_format = '0.00'

                feuille.cell(
                    row=row,
                    column=8
                ).number_format = '0.00'

            # ------------------------------------------------
            # LARGEURS
            # ------------------------------------------------

            largeurs = [
                12,
                24,
                12,
                25,
                20,
                15,
                25,
                22,
                32
            ]

            for i, largeur in enumerate(
                largeurs,
                1
            ):

                feuille.column_dimensions[
                    get_column_letter(i)
                ].width = largeur

            # ------------------------------------------------
            # BORDURES
            # ------------------------------------------------

            bordure = Border(
                left=Side(style="thin", color="B7B7B7"),
                right=Side(style="thin", color="B7B7B7"),
                top=Side(style="thin", color="B7B7B7"),
                bottom=Side(style="thin", color="B7B7B7")
            )

            for row in feuille.iter_rows(
                min_row=4,
                max_row=max(4, ligne - 1),
                min_col=1,
                max_col=9
            ):

                for cellule in row:
                    cellule.border = bordure

            # ------------------------------------------------
            # GEL DE L'EN-TÊTE
            # ------------------------------------------------

            feuille.freeze_panes = "A5"

            # ------------------------------------------------
            # FILTRES
            # ------------------------------------------------

            if ligne > 5:
                feuille.auto_filter.ref = (
                    f"A4:I{ligne - 1}"
                )

            classeur.save(fichier)

            messagebox.showinfo(
                "Export Excel",
                "Fichier Excel professionnel créé avec succès."
            )

        except PermissionError:

            messagebox.showerror(
                "Erreur",
                "Impossible d'enregistrer le fichier.\n"
                "Vérifiez qu'il n'est pas déjà ouvert dans Excel."
            )

        except Exception as erreur:

            messagebox.showerror(
                "Erreur",
                f"Impossible de créer le fichier Excel :\n{erreur}"
            )

    # ========================================================
    # RAPPORT AUTOMATIQUE
    # ========================================================

    def generer_rapport(self):

        if not self.projets:

            messagebox.showwarning(
                "Attention",
                "Aucun projet disponible."
            )

            return

        projet = self.projets[-1]

        fichier = filedialog.asksaveasfilename(
            title="Enregistrer le rapport",
            defaultextension=".txt",
            filetypes=[
                ("Fichier texte", "*.txt")
            ],
            initialfile=(
                f"rapport_projet_{projet['numero']}.txt"
            )
        )

        if not fichier:
            return

        try:

            date_analyse = datetime.now().strftime(
                "%d/%m/%Y %H:%M"
            )

            with open(
                fichier,
                "w",
                encoding="utf-8"
            ) as rapport:

                rapport.write(
                    "=" * 65 + "\n"
                )

                rapport.write(
                    "        RAPPORT D'ANALYSE FINANCIÈRE\n"
                )

                rapport.write(
                    "                 VAN & TRI\n"
                )

                rapport.write(
                    "=" * 65 + "\n\n"
                )

                rapport.write(
                    f"Date d'analyse : {date_analyse}\n"
                )

                rapport.write(
                    f"Projet : {projet['numero']}\n\n"
                )

                rapport.write(
                    "CARACTÉRISTIQUES DU PROJET\n"
                )

                rapport.write(
                    "-" * 65 + "\n"
                )

                rapport.write(
                    f"Investissement initial : "
                    f"{formater_nombre(projet['investissement'])} FCFA\n"
                )

                rapport.write(
                    f"Durée : {projet['annees']} ans\n"
                )

                rapport.write(
                    f"Taux d'actualisation : "
                    f"{formater_pourcentage(projet['taux'])}\n\n"
                )

                rapport.write(
                    "RÉSULTATS FINANCIERS\n"
                )

                rapport.write(
                    "-" * 65 + "\n"
                )

                rapport.write(
                    f"VAN : "
                    f"{formater_nombre(projet['van'])} FCFA\n"
                )

                rapport.write(
                    f"TRI : "
                    f"{formater_pourcentage(projet['tri'])}\n"
                )

                rapport.write(
                    f"Indice de rentabilité : "
                    f"{formater_nombre(projet['indice'])}\n"
                )

                if projet["recuperation"] is not None:

                    rapport.write(
                        f"Délai de récupération : "
                        f"{formater_nombre(projet['recuperation'])} ans\n"
                    )

                else:

                    rapport.write(
                        "Délai de récupération : Non récupéré\n"
                    )

                rapport.write(
                    "\nDÉCISION\n"
                )

                rapport.write(
                    "-" * 65 + "\n"
                )

                rapport.write(
                    projet["decision"] + "\n"
                )

                rapport.write(
                    "\nVÉRIFICATION DU TRI\n"
                )

                rapport.write(
                    "-" * 65 + "\n"
                )

                rapport.write(
                    f"VAN au TRI : "
                    f"{projet['van_verification']:.8f}\n"
                )

                if projet["tri_precis"]:

                    rapport.write(
                        "Résultat : TRI suffisamment précis.\n"
                    )

                else:

                    rapport.write(
                        "Résultat : vérification supplémentaire recommandée.\n"
                    )

                if len(projet["tris"]) > 1:

                    rapport.write(
                        "\nAVERTISSEMENT\n"
                    )

                    rapport.write(
                        "Plusieurs TRI ont été détectés : "
                        + ", ".join(
                            formater_pourcentage(t)
                            for t in projet["tris"]
                        )
                        + "\n"
                    )

                    rapport.write(
                        "La décision doit être interprétée avec prudence.\n"
                    )

                rapport.write(
                    "\nFLUX FINANCIERS\n"
                )

                rapport.write(
                    "-" * 65 + "\n"
                )

                for annee, flux in enumerate(
                    projet["flux"]
                ):

                    if annee == 0:
                        rapport.write(
                            f"Date 0 : "
                            f"{formater_nombre(flux)} FCFA\n"
                        )
                    else:
                        rapport.write(
                            f"Année {annee} : "
                            f"{formater_nombre(flux)} FCFA\n"
                        )

                rapport.write(
                    "\n" + "=" * 65 + "\n"
                )

                rapport.write(
                    "Fin du rapport.\n"
                )

            messagebox.showinfo(
                "Rapport",
                "Rapport créé avec succès."
            )

        except PermissionError:

            messagebox.showerror(
                "Erreur",
                "Impossible d'enregistrer le rapport. "
                "Vérifiez que le fichier n'est pas déjà ouvert."
            )

        except Exception as erreur:

            messagebox.showerror(
                "Erreur",
                f"Impossible de créer le rapport :\n{erreur}"
            )


# ============================================================
# 5. LANCEMENT DE L'APPLICATION
# ============================================================

if __name__ == "__main__":

    fenetre = tk.Tk()

    application = CalculateurFinancier(
        fenetre
    )

    fenetre.mainloop()
